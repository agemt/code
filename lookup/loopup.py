import os
import csv
import shutil
import getpass
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import TableColumn

# ==========================================
# CONFIGURATION
# ==========================================
KEYWORD_Y = "Take"
FORMAT_SHIFT_DATE = datetime(2024, 3, 15)  # Matches "old_peeks" in VBA
PATH_REGISTRY_COLUMNS = ["Path", "Engine", "Date_Tested", "Date_Added", "Added_By"]


# ==========================================
# HELPER FUNCTIONS
# ==========================================
def _create_backup(filepath):
    """Creates a backup, retaining only the 2 latest."""
    if not os.path.exists(filepath): return None
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = f"{filepath}.{timestamp}.bak"
    shutil.copy2(filepath, backup_path)
    
    directory = os.path.dirname(filepath) or "."
    filename = os.path.basename(filepath)
    backups = sorted([os.path.join(directory, f) for f in os.listdir(directory) 
                      if f.startswith(f"{filename}.") and f.endswith(".bak")], 
                     key=os.path.getmtime)
    for old_backup in backups[:-2]:
        try: os.remove(old_backup)
        except OSError: pass
    return backup_path

def _expand_excel_table(ws):
    """Expands Excel ListObject bounds and keeps table metadata in sync."""
    if not ws.tables:
        return

    table = list(ws.tables.values())[0]
    min_col, min_row, _max_col, _max_row = range_boundaries(table.ref)
    new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
    table.ref = new_ref

    # openpyxl does not rebuild table column metadata when only ref is widened.
    # Excel treats that mismatch as a corrupted ListObject on open.
    header_row = min_row
    header_names = []
    seen_names = {}

    for col_idx in range(min_col, ws.max_column + 1):
        raw_name = ws.cell(row=header_row, column=col_idx).value
        base_name = str(raw_name).strip() if raw_name not in {None, ""} else f"Column{col_idx}"
        unique_name = base_name
        if unique_name in seen_names:
            seen_names[unique_name] += 1
            unique_name = f"{unique_name}_{seen_names[unique_name]}"
            ws.cell(row=header_row, column=col_idx, value=unique_name)
        else:
            seen_names[unique_name] = 1
        header_names.append(unique_name)

    table.tableColumns = [
        TableColumn(id=offset, name=header_name)
        for offset, header_name in enumerate(header_names, start=1)
    ]

    if table.autoFilter is not None:
        table.autoFilter.ref = new_ref

def _clean_value(raw_val):
    """Handles commas and numeric conversions."""
    if pd.isna(raw_val): return None
    val_str = str(raw_val).strip()
    if val_str.lower() in ["nan", "none", ""]: return None
    val_str = val_str.replace(',', '.')
    try:
        numeric_val = float(val_str)
        return int(numeric_val) if numeric_val.is_integer() else numeric_val
    except ValueError: return val_str


def _find_header_index(header_map, aliases):
    """Finds a column index by normalized aliases from an Excel header map."""
    normalized = {
        str(name).strip().lower().replace(" ", "_").replace("-", "_"): idx
        for name, idx in header_map.items()
    }
    for alias in aliases:
        key = alias.strip().lower().replace(" ", "_").replace("-", "_")
        if key in normalized:
            return normalized[key]
    return None


def _normalize_date_key(value):
    """Normalizes date-like values to a stable key for uniqueness checks."""
    raw = str(value).strip()
    if not raw or raw.lower() in {"nan", "none"}:
        return ""

    parsed = pd.to_datetime(raw, errors="coerce")
    if pd.notna(parsed):
        return parsed.strftime("%Y-%m-%d %H:%M:%S")
    return raw


def _default_paths_registry_path():
    """Returns default registry path, preferring paths.xlsx if present."""
    lookup_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(lookup_dir, "paths.xlsx")
    txt_path = os.path.join(lookup_dir, "paths.txt")
    return xlsx_path if os.path.exists(xlsx_path) else txt_path


def _get_windows_user():
    """Best-effort resolution of the active Windows username."""
    return (
        os.environ.get("USERNAME")
        or os.environ.get("USER")
        or getpass.getuser()
        or "unknown"
    )


def _fallback_engine_name(path):
    """Uses the parent folder name first when engine metadata is missing."""
    normalized_path = os.path.normpath(str(path or "").strip())
    parent_name = os.path.basename(os.path.dirname(normalized_path)).strip()
    if parent_name:
        return parent_name
    return os.path.splitext(os.path.basename(normalized_path))[0].strip() or "Unknown"


def _describe_path(path):
    """Returns a compact parent/filename label for progress messages."""
    normalized_path = os.path.normpath(str(path or "").strip())
    filename = os.path.basename(normalized_path)
    parent_name = os.path.basename(os.path.dirname(normalized_path)).strip()
    if parent_name and filename:
        return f"{parent_name}\\{filename}"
    return filename or normalized_path


def _normalize_paths_df(df):
    """Normalizes registry columns while preserving any extra metadata columns."""
    if df is None or df.empty:
        return pd.DataFrame(columns=PATH_REGISTRY_COLUMNS)

    # Alias known variants into canonical names.
    rename_map = {}
    for col in df.columns:
        if not isinstance(col, str):
            continue
        compact = col.strip().lower().replace("-", "_").replace(" ", "_")
        if compact in {"path", "file_path"}:
            rename_map[col] = "Path"
        elif compact in {"engine", "engine_name"}:
            rename_map[col] = "Engine"
        elif compact in {"date_tested", "test_date"}:
            rename_map[col] = "Date_Tested"
        elif compact in {"date_added", "added_date"}:
            rename_map[col] = "Date_Added"
        elif compact in {"added_by", "who_added", "user"}:
            rename_map[col] = "Added_By"

    normalized = df.rename(columns=rename_map).copy()

    # Backward compatibility: plain text file may contain only paths with no header.
    if "Path" not in normalized.columns and len(normalized.columns) == 1:
        normalized = normalized.rename(columns={normalized.columns[0]: "Path"})

    for col in PATH_REGISTRY_COLUMNS:
        if col not in normalized.columns:
            normalized[col] = ""

    return normalized


def _extract_latest_test_date(run_path):
    """Extracts latest matching test date from a run file using lookup coordinates."""
    if not run_path or not os.path.exists(run_path):
        return ""

    df = load_raw_grid(run_path)
    if df is None or df.empty:
        return ""

    coords = _get_lookup_coords(run_path, df=df)
    date_candidates = []
    raw_fallback = []

    total_cols = len(df.columns)
    for data_col in range(2, total_cols):
        try:
            test_val = str(df.iloc[coords["test_row"], data_col]).strip()
        except Exception:
            continue

        if KEYWORD_Y.lower() not in test_val.lower():
            continue

        try:
            raw_date_val = df.iloc[coords["date_row"], data_col]
        except Exception:
            continue

        if pd.isna(raw_date_val):
            continue

        raw_date_str = str(raw_date_val).strip()
        if raw_date_str:
            raw_fallback.append(raw_date_str)

        parsed = pd.to_datetime(raw_date_val, errors="coerce")
        if pd.notna(parsed):
            date_candidates.append(parsed)

    if date_candidates:
        latest = max(date_candidates)
        return latest.strftime("%Y-%m-%d %H:%M:%S")

    return raw_fallback[-1] if raw_fallback else ""


def preview_run_file(run_path):
    """Returns a summary preview (engine, matching take tests, tested dates) for one run file."""
    normalized_path = str(run_path or "").strip()
    if not normalized_path:
        return {"ok": False, "message": "Path is empty."}

    normalized_path = os.path.normpath(normalized_path)
    if not os.path.exists(normalized_path):
        return {"ok": False, "message": f"File not found: {normalized_path}"}

    df = load_raw_grid(normalized_path)
    if df is None or df.empty:
        return {"ok": False, "message": "Unable to parse file content."}

    coords = _get_lookup_coords(normalized_path, df=df)
    try:
        engine_val = str(df.iloc[coords["engine_row"], coords["engine_col"]]).strip()
    except Exception:
        engine_val = ""

    engine_detected = True
    if (
        not engine_val
        or engine_val.lower() in {"nan", "unknown"}
        or KEYWORD_Y.lower() in engine_val.lower()
        or engine_val.startswith("#")
    ):
        engine_detected = False
        engine_val = ""

    take_tests = []
    tested_dates = []
    for data_col in range(2, len(df.columns)):
        try:
            test_val = str(df.iloc[coords["test_row"], data_col]).strip()
        except Exception:
            continue

        if KEYWORD_Y.lower() not in test_val.lower():
            continue

        take_tests.append(test_val)
        try:
            raw_date_val = df.iloc[coords["date_row"], data_col]
            if pd.notna(raw_date_val):
                tested_dates.append(str(raw_date_val).strip())
        except Exception:
            pass

    return {
        "ok": True,
        "path": normalized_path,
        "engine": engine_val,
        "engine_detected": engine_detected,
        "take_tests": take_tests,
        "tested_dates": tested_dates,
        "message": f"Found {len(take_tests)} matching '{KEYWORD_Y}' test columns.",
    }


def read_paths_registry(paths_registry=None, latest_first=True):
    """Reads paths registry from xlsx/txt and returns a normalized DataFrame."""
    registry_path = paths_registry or _default_paths_registry_path()

    if not os.path.exists(registry_path) or os.path.getsize(registry_path) == 0:
        return pd.DataFrame(columns=PATH_REGISTRY_COLUMNS)

    ext = os.path.splitext(registry_path)[1].lower()
    try:
        if ext in {".xlsx", ".xls"}:
            raw = pd.read_excel(registry_path)
        else:
            raw = pd.read_csv(registry_path, encoding="utf-8-sig")
    except Exception:
        # Last resort for old txt files that were just newline separated paths.
        with open(registry_path, "r", encoding="utf-8-sig") as f:
            rows = [line.strip() for line in f if line.strip()]
        raw = pd.DataFrame({"Path": rows})

    normalized = _normalize_paths_df(raw)
    normalized["Path"] = normalized["Path"].astype(str).str.strip()
    normalized = normalized[normalized["Path"] != ""]

    if latest_first and not normalized.empty:
        added_sort = pd.to_datetime(normalized["Date_Added"], errors="coerce")
        normalized = (
            normalized.assign(_added_sort=added_sort)
            .sort_values(by="_added_sort", ascending=False, na_position="last")
            .drop(columns=["_added_sort"])
            .reset_index(drop=True)
        )

    return normalized


def _write_paths_registry(df, paths_registry=None):
    """Persists normalized paths registry to xlsx/txt depending on destination extension."""
    registry_path = paths_registry or _default_paths_registry_path()
    ext = os.path.splitext(registry_path)[1].lower()

    out_df = _normalize_paths_df(df)
    out_df = out_df[PATH_REGISTRY_COLUMNS + [c for c in out_df.columns if c not in PATH_REGISTRY_COLUMNS]]

    if ext in {".xlsx", ".xls"}:
        out_df.to_excel(registry_path, index=False)
    else:
        out_df.to_csv(registry_path, index=False, encoding="utf-8-sig")


def add_path_registry_entry(new_path, paths_registry=None, added_by=None, engine_name=None):
    """Adds a new path with metadata (date tested, date added, windows user)."""
    normalized_path = str(new_path or "").strip()
    if not normalized_path:
        return {"ok": False, "message": "Path is empty."}

    normalized_path = os.path.normpath(normalized_path)
    current_df = read_paths_registry(paths_registry=paths_registry, latest_first=False)

    existing = current_df["Path"].astype(str).str.lower().str.strip().tolist() if not current_df.empty else []
    if normalized_path.lower() in existing:
        return {"ok": False, "message": "Path already exists in registry."}

    preview = preview_run_file(normalized_path)
    detected_engine = ""
    if preview.get("ok") and preview.get("engine_detected"):
        detected_engine = str(preview.get("engine") or "").strip()

    selected_engine = str(engine_name or "").strip() or detected_engine or _fallback_engine_name(normalized_path)

    record = {
        "Path": normalized_path,
        "Engine": selected_engine,
        "Date_Tested": _extract_latest_test_date(normalized_path),
        "Date_Added": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Added_By": (added_by or _get_windows_user()),
    }

    updated = pd.concat([current_df, pd.DataFrame([record])], ignore_index=True)
    _write_paths_registry(updated, paths_registry=paths_registry)
    return {"ok": True, "message": "Path added.", "record": record}


def get_latest_paths(limit=100, paths_registry=None):
    """Returns latest registry rows as records sorted by Date_Added descending."""
    df = read_paths_registry(paths_registry=paths_registry, latest_first=True)
    if limit and limit > 0:
        df = df.head(limit)
    return df.to_dict("records")

def load_raw_grid(path):
    """Loads file as a raw grid (no headers) for direct spatial coordinates."""
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext in ['.xlsx', '.xls']:
            engine = 'openpyxl' if ext == '.xlsx' else 'xlrd'
            return pd.read_excel(path, engine=engine, header=None)
        
        for enc in ['utf-8-sig', 'utf-16', 'latin1']:
            try:
                return pd.read_csv(path, sep=None, engine='python', header=None, 
                                   quoting=csv.QUOTE_NONE, escapechar='\\', 
                                   on_bad_lines='skip', encoding=enc)
            except Exception: continue
    except Exception: pass
    return None

def _get_lookup_coords(filepath, df=None):
    """Determines metadata coordinates and supports optional title row files."""
    if df is not None and not df.empty and len(df.columns) > 2:
        max_scan_rows = min(8, len(df))
        detected_test_row = None
        best_take_hits = 0

        for r_idx in range(max_scan_rows):
            take_hits = 0
            for c_idx in range(2, len(df.columns)):
                raw_val = df.iloc[r_idx, c_idx]
                if pd.isna(raw_val):
                    continue
                if KEYWORD_Y.lower() in str(raw_val).lower():
                    take_hits += 1

            if take_hits > best_take_hits:
                best_take_hits = take_hits
                detected_test_row = r_idx

        if detected_test_row is not None and best_take_hits > 0:
            date_row = min(detected_test_row + 1, len(df) - 1)
            engine_row = max(detected_test_row - 1, 0)
            return {
                "engine_row": engine_row,
                "engine_col": 2,
                "test_row": detected_test_row,
                "date_row": date_row,
            }

    # Historical fallback defaults.
    file_ctime = datetime.fromtimestamp(os.path.getctime(filepath))
    if file_ctime > FORMAT_SHIFT_DATE:
        return {"engine_row": 0, "engine_col": 2, "test_row": 1, "date_row": 2}
    return {"engine_row": 0, "engine_col": 2, "test_row": 1, "date_row": 2}

# ==========================================
# FUNCTION 1: INGESTER (MIMICKING VBA LOGIC)
# ==========================================
def ingest_new_runs(new_runs, all_params, master_excel, engine_overrides=None):
    """
    new_runs: List of file paths
    """
    total_steps = len(new_runs) + 2
    current_step = 1

    yield {"progress": current_step, "total": total_steps, "message": "Backing up master database..."}
    _create_backup(master_excel)
    current_step += 1

    wb = openpyxl.load_workbook(master_excel)
    ws = wb.active
    
    # Maps parameter names to their target column index in Master Excel
    header_map = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1) if cell.value}
    engine_col_idx = _find_header_index(header_map, ["Engine", "Engine_Name"])
    date_col_idx = _find_header_index(header_map, ["Date_Tested", "Date Tested", "Date", "Full_date"])
    perf_col_idx = _find_header_index(header_map, ["Perf. Point", "Perf Point", "Perf_Point", "Perf"])

    existing_date_keys = set()
    if date_col_idx:
        for r_idx in range(2, ws.max_row + 1):
            date_cell_val = ws.cell(row=r_idx, column=date_col_idx).value
            key = _normalize_date_key(date_cell_val)
            if key:
                existing_date_keys.add(key)

    for idx, path in enumerate(new_runs, start=1):
        path_label = _describe_path(path)
        yield {"progress": current_step, "total": total_steps, "message": f"Processing [{idx}/{len(new_runs)}]: {path_label}"}
        current_step += 1

        df = load_raw_grid(path)
        if df is None or df.empty: continue

        coords = _get_lookup_coords(path, df=df)
        
        # VBA: engine = sourcerange.Cells(1, 3).Value
        try:
            engine_val = str(df.iloc[coords["engine_row"], coords["engine_col"]]).strip()
        except IndexError:
            engine_val = "Unknown"

        if (
            engine_val.lower() in {"nan", "", "unknown"}
            or KEYWORD_Y.lower() in engine_val.lower()
            or engine_val.startswith("#")
        ):
            engine_val = _fallback_engine_name(path)

        if engine_overrides and path in engine_overrides:
            manual_engine = str(engine_overrides.get(path) or "").strip()
            if manual_engine:
                engine_val = manual_engine

        # Walk horizontally across data columns (Starts at Col C -> Pandas index 2)
        total_cols = len(df.columns)
        for data_col in range(2, total_cols):
            
            # VBA: test = sourcerange.Cells(2, i).Value
            test_val = str(df.iloc[coords["test_row"], data_col]).strip()
            
            # VBA: If (InStr(test, "Take") > 0) Then
            if KEYWORD_Y.lower() in test_val.lower():
                
                # VBA: date_tested = sourcerange.Cells(3, i).Value
                date_val = str(df.iloc[coords["date_row"], data_col]).strip()
                date_key = _normalize_date_key(date_val)

                # Date is the unique key; duplicate dates are skipped.
                if date_col_idx and date_key and date_key in existing_date_keys:
                    continue
                
                # VBA: target_range_take.Rows(2).EntireRow.Insert
                ws.insert_rows(2)
                
                # Write Identifiers (Engine, Date, Test)
                if engine_col_idx: ws.cell(row=2, column=engine_col_idx, value=engine_val)
                if date_col_idx: ws.cell(row=2, column=date_col_idx, value=date_val)
                if perf_col_idx: ws.cell(row=2, column=perf_col_idx, value=test_val)

                if date_col_idx and date_key:
                    existing_date_keys.add(date_key)

                # Walk vertically down Column A to find parameters
                first_col = df.iloc[:, 0].astype(str).str.strip().tolist()
                total_rows = len(df)
                
                for param_row in range(total_rows):
                    param_name = first_col[param_row]
                    
                    # Replaces the massive VBA "Select Case" block
                    if param_name in all_params and param_name in header_map:
                        
                        # VBA: sourcerange.Cells(j, i).Value
                        raw_data_val = df.iloc[param_row, data_col]
                        
                        target_col_idx = header_map[param_name]
                        ws.cell(row=2, column=target_col_idx, value=_clean_value(raw_data_val))

    yield {"progress": total_steps - 1, "total": total_steps, "message": "Formatting table bounds..."}
    _expand_excel_table(ws)
    wb.save(master_excel)
    yield {"progress": total_steps, "total": total_steps, "message": "Ingestion complete."}

# ==========================================
# FUNCTION 2: SINGLE PARAMETER RETROACTIVE LOOKUP
# ==========================================
def retroactive_parameter_update(new_param_str, paths_excel, params_txt, master_excel):
    # Determine header name
    header = new_param_str.split("=")[0].strip() if "=" in new_param_str else new_param_str.strip()
    
    # Abort if parameter already tracked
    if os.path.exists(params_txt):
        with open(params_txt, 'r', encoding='utf-8-sig') as f:
            existing_params = {line.strip() for line in f}
            if new_param_str in existing_params or header in existing_params:
                yield {"progress": 100, "total": 100, "message": f"ABORTED: '{header}' already exists in params.txt."}
                return

    yield {"progress": 1, "total": 100, "message": "Initializing parameter update..."}
        
    _create_backup(master_excel)
    
    wb = openpyxl.load_workbook(master_excel)
    ws = wb.active
    
    # Append to right-most column
    new_col_idx = ws.max_column + 1
    ws.cell(row=1, column=new_col_idx, value=header)
    
    if "=" in new_param_str:
        # Excel Formula Injection
        formula = new_param_str.split("=", 1)[1].strip()
        if not formula.startswith("="): formula = "=" + formula
        yield {"progress": 50, "total": 100, "message": f"Injecting formula for {header}..."}
        
        for r_idx in range(2, ws.max_row + 1):
            ws.cell(row=r_idx, column=new_col_idx, value=formula)
    else:
        # Map target rows strictly by unique tested date, with tolerant header matching.
        header_map = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1) if cell.value}
        normalized_header_map = {
            str(k).strip().lower().replace(" ", "_").replace("-", "_"): v
            for k, v in header_map.items()
        }

        preferred_date_keys = [
            "date_tested",
            "date_test",
            "date_tested_utc",
            "date",
            "full_date",
            "date_tested_local",
        ]

        date_col = None
        for key in preferred_date_keys:
            if key in normalized_header_map:
                date_col = normalized_header_map[key]
                break

        # Last-resort fallback: any header containing both "date" and "test".
        if not date_col:
            for norm_key, idx in normalized_header_map.items():
                if "date" in norm_key and "test" in norm_key:
                    date_col = idx
                    break

        if not date_col:
            yield {
                "progress": 100,
                "total": 100,
                "message": "Error: tested date column missing in Master (expected aliases like Date_Tested/Date Tested/Date).",
            }
            return
            
        master_row_map = {}
        for r_idx in range(2, ws.max_row + 1):
            date_in_master = ws.cell(row=r_idx, column=date_col).value
            date_key = _normalize_date_key(date_in_master)
            if date_key:
                master_row_map[date_key] = r_idx

        # Scan files
        config_df = read_paths_registry(paths_registry=paths_excel, latest_first=False)
        if config_df.empty or 'Path' not in config_df.columns:
            yield {"progress": 100, "total": 100, "message": "Error: no paths available in registry."}
            return
        total_files = len(config_df)
        
        for idx, row in config_df.iterrows():
            path = str(row['Path']).strip()
            yield {"progress": 10 + idx, "total": total_files + 20, "message": f"Scanning: {_describe_path(path)}"}
            
            df = load_raw_grid(path)
            if df is None or df.empty: continue
            
            coords = _get_lookup_coords(path, df=df)
            
            # Find the specific row for this new parameter
            first_col = df.iloc[:, 0].astype(str).str.strip().tolist()
            if header not in first_col: continue
            target_param_row = first_col.index(header)

            total_cols = len(df.columns)
            for data_col in range(2, total_cols):
                test_val = str(df.iloc[coords["test_row"], data_col]).strip()
                
                if KEYWORD_Y.lower() in test_val.lower():
                    date_val = str(df.iloc[coords["date_row"], data_col]).strip()
                    date_key = _normalize_date_key(date_val)

                    # Match unique Date_Tested to find correct row in Master
                    if date_key in master_row_map:
                        ws_row = master_row_map[date_key]
                        raw_data_val = df.iloc[target_param_row, data_col]

                        ws.cell(row=ws_row, column=new_col_idx, value=_clean_value(raw_data_val))

    yield {"progress": 99, "total": 100, "message": "Expanding table boundaries..."}
    _expand_excel_table(ws)
    wb.save(master_excel)

    # Append to params only after successful workbook save.
    params_exists = os.path.exists(params_txt)
    if params_exists and os.path.getsize(params_txt) > 0:
        with open(params_txt, 'a', encoding='utf-8-sig') as f:
            f.write(f"\n{new_param_str}")
    else:
        with open(params_txt, 'w', encoding='utf-8-sig') as f:
            f.write(new_param_str)
    
    yield {"progress": 100, "total": 100, "message": f"Parameter '{header}' successfully added."}