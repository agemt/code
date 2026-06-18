
import os
import csv
import shutil
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

# ==========================================
# CONFIGURATION
# ==========================================
KEYWORD_Y = "Take"

# ==========================================
# HELPER FUNCTIONS
# ==========================================
def _create_backup(filepath):
    """Creates a timestamped backup of the master file, keeping only the 2 latest."""
    if not os.path.exists(filepath):
        return None

    # 1. Create the new backup
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = f"{filepath}.{timestamp}.bak"
    shutil.copy2(filepath, backup_path)

    # 2. Cleanup old backups
    directory = os.path.dirname(filepath) or "."
    filename = os.path.basename(filepath)
    
    # Find all backup files for this specific master file
    backups = [
        os.path.join(directory, f) for f in os.listdir(directory)
        if f.startswith(f"{filename}.") and f.endswith(".bak")
    ]
    
    # Sort backups by modification time (oldest first)
    backups.sort(key=os.path.getmtime)
    
    # Delete all but the 2 most recent backups
    for old_backup in backups[:-2]:
        try:
            os.remove(old_backup)
        except OSError:
            pass # Failsafe if file is locked or already deleted
            
    return backup_path

def _expand_excel_table(ws):
    """Finds the first ListObject (Table) and expands its boundaries to fit all data."""
    if not ws.tables:
        return
    
    table = list(ws.tables.values())[0]
    max_row = ws.max_row
    max_col = ws.max_column
    
    top_left = table.ref.split(':')[0]
    bottom_right = f"{get_column_letter(max_col)}{max_row}"
    table.ref = f"{top_left}:{bottom_right}"

def _append_to_txt(filepath, new_items):
    """Appends unique items to a text file."""
    existing = set()
    if os.path.exists(filepath):
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            existing = {line.strip() for line in f if line.strip()}
            
    with open(filepath, 'a', encoding='utf-8-sig') as f:
        for item in new_items:
            if item not in existing:
                f.write(f"\n{item}")

def _append_to_config(filepath, rows_to_add):
    """Appends new runs to the paths configuration Excel."""
    if not os.path.exists(filepath):
        pd.DataFrame(rows_to_add, columns=["Path", "Lookup", "Date"]).to_excel(filepath, index=False)
        return

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    for row in rows_to_add:
        ws.append(row)
    _expand_excel_table(ws)
    wb.save(filepath)

def load_data(path):
    """Robustly reads Excel or CSV files with multiple encoding fallbacks."""
    ext = os.path.splitext(path)[1].lower()

    if ext == '.xlsx':
        try: return pd.read_excel(path, engine='openpyxl'), "Excel (.xlsx)"
        except Exception: pass
    if ext == '.xls':
        try: return pd.read_excel(path, engine='xlrd'), "Excel (.xls)"
        except Exception: pass

    try:
        df = pd.read_csv(path, sep=None, engine='python', quoting=csv.QUOTE_NONE, escapechar='\\', on_bad_lines='skip', encoding='utf-8-sig')
        return df, "CSV/TXT (UTF-8)"
    except Exception: pass

    try:
        df = pd.read_csv(path, sep=None, engine='python', quoting=csv.QUOTE_NONE, escapechar='\\', on_bad_lines='skip', encoding='utf-16')
        return df, "CSV/TXT (UTF-16)"
    except Exception: pass

    try:
        df = pd.read_csv(path, sep=None, engine='python', quoting=csv.QUOTE_NONE, escapechar='\\', on_bad_lines='skip', encoding='latin1', encoding_errors='ignore')
        return df, "CSV/TXT (Latin1 Fallback)"
    except Exception:
        return None, "Failed to Parse"

def _clean_value(raw_val):
    """Applies your original comma/dot parsing and numeric checking."""
    if pd.isna(raw_val): return None
    val_str = str(raw_val).strip()
    if val_str.lower() in ["nan", "none", ""]: return None
    
    val_str = val_str.replace(',', '.')
    try:
        numeric_val = float(val_str)
        return int(numeric_val) if numeric_val.is_integer() else numeric_val
    except ValueError:
        return val_str

# ==========================================
# FUNCTION 1: INGESTER
# ==========================================
def ingest_new_runs(new_runs, all_params, paths_excel, params_txt, master_excel):
    """
    new_runs: List of tuples -> [(path, lookup_row, date_row), ...]
    all_params: List of strings (The master list of parameters to extract)
    """
    total_steps = len(new_runs) + 3
    current_step = 1

    yield {"progress": current_step, "total": total_steps, "message": "Updating configurations..."}
    _append_to_config(paths_excel, new_runs)
    current_step += 1

    yield {"progress": current_step, "total": total_steps, "message": "Creating backup of master database..."}
    _create_backup(master_excel)
    current_step += 1

    wb = openpyxl.load_workbook(master_excel)
    ws = wb.active
    
    # 1-based index mapping for openpyxl inserts
    header_map = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1) if cell.value}
    
    for idx, run in enumerate(new_runs, start=1):
        path, lookup_row_val, date_row_val = run
        filename = os.path.basename(path)
        yield {"progress": current_step, "total": total_steps, "message": f"Processing file [{idx}/{len(new_runs)}]: {filename}"}
        current_step += 1

        df, format_type = load_data(path)
        if df is None or df.empty:
            yield {"progress": current_step, "total": total_steps, "message": f"  ❌ Skip: Could not read {filename}"}
            continue

        try:
            row_x_idx = int(float(lookup_row_val))
            date_row_idx = int(float(date_row_val))
        except ValueError:
            continue

        if row_x_idx >= len(df) or date_row_idx >= len(df):
            continue

        row_x = df.iloc[row_x_idx]
        matching_cols = [col for col in df.columns if KEYWORD_Y.lower() in str(row_x[col]).lower()]

        if not matching_cols:
            continue

        first_col_str = df.iloc[:, 0].astype(str).str.strip().tolist()
        param_row_map = {val: i for i, val in enumerate(first_col_str)}
        parent = os.path.basename(os.path.dirname(path))

        for col in matching_cols:
            matched_value = str(row_x[col]).replace(',', '.') if not pd.isna(row_x[col]) else None
            
            raw_date = df.iloc[date_row_idx, df.columns.get_loc(col)]
            try:
                date_value = pd.to_datetime(raw_date).strftime('%Y-%m-%d %H:%M:%S') if not pd.isna(raw_date) else None
            except Exception:
                date_value = str(raw_date) if not pd.isna(raw_date) else None

            # Build row
            new_row = [None] * len(header_map)
            if 'Engine' in header_map: new_row[header_map['Engine'] - 1] = parent
            if 'Date_Tested' in header_map: new_row[header_map['Date_Tested'] - 1] = date_value
            if 'Perf. Point' in header_map: new_row[header_map['Perf. Point'] - 1] = matched_value

            for p in all_params:
                clean_p = str(p).strip()
                if clean_p in param_row_map:
                    row_idx = param_row_map[clean_p]
                    raw_val = df.iloc[row_idx, df.columns.get_loc(col)]
                    cleaned_val = _clean_value(raw_val)
                    
                    if clean_p in header_map:
                        new_row[header_map[clean_p] - 1] = cleaned_val

            ws.append(new_row)

    yield {"progress": current_step, "total": total_steps, "message": "Saving and expanding table bounds..."}
    _expand_excel_table(ws)
    wb.save(master_excel)
    yield {"progress": total_steps, "total": total_steps, "message": "Ingestion complete."}

# ==========================================
# FUNCTION 2: RETROACTIVE UPDATER
# ==========================================
def retroactive_parameter_update(new_param_str, paths_excel, params_txt, master_excel):
    """
    new_param_str: e.g. "Fuel_Flow" or "Thrust_Ratio=[@[Gross_Thrust]];[@[Net_Thrust]]"
    """
    yield {"progress": 1, "total": 100, "message": "Initializing parameter update..."}
    
    _append_to_txt(params_txt, [new_param_str])
    _create_backup(master_excel)
    
    wb = openpyxl.load_workbook(master_excel)
    ws = wb.active
    
    new_col_idx = ws.max_column + 1
    max_row = ws.max_row
    
    if "=" in new_param_str:
        # HEADER=FORMULA
        header, formula = new_param_str.split("=", 1)
        header = header.strip()
        formula = formula.strip()
        if not formula.startswith("="): formula = "=" + formula
            
        yield {"progress": 50, "total": 100, "message": f"Injecting formula for {header}..."}
        ws.cell(row=1, column=new_col_idx, value=header)
        for row_idx in range(2, max_row + 1):
            ws.cell(row=row_idx, column=new_col_idx, value=formula)
            
    else:
        # RAW CHANNEL
        header = new_param_str.strip()
        ws.cell(row=1, column=new_col_idx, value=header)
        
        config_df = pd.read_excel(paths_excel)
        total_files = len(config_df)
        
        yield {"progress": 10, "total": total_files + 20, "message": f"Mapping master database rows..."}
        
        # Map existing database rows to allow targeting specific cells
        # format: { (Engine, Date_Tested, Perf Point): Row_Index }
        header_map = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1) if cell.value}
        master_row_map = {}
        
        eng_col = header_map.get('Engine')
        date_col = header_map.get('Date_Tested')
        perf_col = header_map.get('Perf. Point')
        
        if all([eng_col, date_col, perf_col]):
            for r_idx in range(2, max_row + 1):
                e = str(ws.cell(row=r_idx, column=eng_col).value)
                d = str(ws.cell(row=r_idx, column=date_col).value)
                p = str(ws.cell(row=r_idx, column=perf_col).value)
                master_row_map[(e, d, p)] = r_idx

        # Scan files
        for idx, row in config_df.iterrows():
            path = str(row['Path']).strip()
            yield {"progress": 10 + idx, "total": total_files + 20, "message": f"Scanning: {os.path.basename(path)}"}
            
            df, _ = load_data(path)
            if df is None or df.empty: continue

            try:
                row_x_idx = int(float(row['Lookup']))
                date_row_idx = int(float(row['Date']))
            except ValueError: continue

            if row_x_idx >= len(df) or date_row_idx >= len(df): continue

            row_x = df.iloc[row_x_idx]
            matching_cols = [c for c in df.columns if KEYWORD_Y.lower() in str(row_x[c]).lower()]
            if not matching_cols: continue

            first_col_str = df.iloc[:, 0].astype(str).str.strip().tolist()
            param_row_map = {val: i for i, val in enumerate(first_col_str)}
            parent = os.path.basename(os.path.dirname(path))

            if header in param_row_map:
                target_r_idx = param_row_map[header]
                
                for col in matching_cols:
                    matched_value = str(row_x[col]).replace(',', '.') if not pd.isna(row_x[col]) else None
                    raw_date = df.iloc[date_row_idx, df.columns.get_loc(col)]
                    try:
                        date_value = pd.to_datetime(raw_date).strftime('%Y-%m-%d %H:%M:%S') if not pd.isna(raw_date) else None
                    except Exception:
                        date_value = str(raw_date) if not pd.isna(raw_date) else None
                    
                    # Lookup exact row in master Excel
                    key = (parent, str(date_value), str(matched_value))
                    if key in master_row_map:
                        ws_row = master_row_map[key]
                        raw_val = df.iloc[target_r_idx, df.columns.get_loc(col)]
                        cleaned_val = _clean_value(raw_val)
                        
                        ws.cell(row=ws_row, column=new_col_idx, value=cleaned_val)

    yield {"progress": total_files + 19, "total": total_files + 20, "message": "Expanding table boundaries and saving..."}
    _expand_excel_table(ws)
    wb.save(master_excel)
    
    yield {"progress": total_files + 20, "total": total_files + 20, "message": f"Parameter '{header}' successfully added."}
