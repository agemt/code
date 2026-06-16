import re
import csv
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Helper to clean numeric values (strip units, convert to float)
def clean_numeric(val):
    if pd.isna(val) or val is None:
        return np.nan
    val_str = str(val).strip()
    if not val_str or val_str.upper() in ["N/A", "NAN", "NULL", "-"]:
        return np.nan
    # Split by space to strip units like 'G', 'psi', etc.
    parts = val_str.split()
    if parts:
        num_part = parts[0]
        # Remove non-numeric characters except digits, dots, signs
        num_part = re.sub(r'[^\d\.\-\+eE]', '', num_part)
        try:
            return float(num_part)
        except ValueError:
            return np.nan
    return np.nan

# Simple parser to evaluate Excel-like formulas in Python
def evaluate_formula_in_python(df, formula):
    """
    Evaluates a formula like 'MAX([@[W2K3]];0)' or '[@[FPR]]*1.01' on a pandas DataFrame.
    """
    expr = formula
    if expr.startswith("="):
        expr = expr[1:]
        
    # Replace semicolons with commas
    expr = expr.replace(";", ",")
    
    # Replace Excel table placeholders [@[PARAM]] with df['PARAM']
    placeholders = re.findall(r'\[@?\[?([^\]]+)\]?\]', expr)
    for ph in placeholders:
        clean_ph = ph.replace("@", "").strip()
        # Replace the placeholder in the expression
        expr = expr.replace(f"[@[{ph}]]", f"df['{clean_ph}']")
        expr = expr.replace(f"[@{ph}]", f"df['{clean_ph}']")
        expr = expr.replace(f"[{ph}]", f"df['{clean_ph}']")
        
    # Replace basic excel functions with numpy equivalents
    expr = re.sub(r'\bMAX\b', 'np.maximum', expr, flags=re.IGNORECASE)
    expr = re.sub(r'\bMIN\b', 'np.minimum', expr, flags=re.IGNORECASE)
    expr = re.sub(r'\bABS\b', 'np.abs', expr, flags=re.IGNORECASE)
    expr = re.sub(r'\bIF\b', 'np.where', expr, flags=re.IGNORECASE)
    
    # Evaluate context
    eval_dict = {
        "df": df,
        "np": np,
        "pd": pd
    }
    
    try:
        # Check if the columns exist in df
        result = eval(expr, eval_dict)
        return result
    except Exception as e:
        # Fallback to NaN if evaluation fails
        return pd.Series(np.nan, index=df.index)

def parse_raw_file(file_path, params, engine_row_1based, test_row_1based, date_row_1based, custom_engine_name=None, operator_name="Unknown"):
    """
    Parses a single raw file.
    - Delimiter auto-detection.
    - Matches keyword 'Take' in test row.
    - Extracts engine name, test scope, date, and parameter values.
    """
    # 0-indexed coordinates
    engine_idx = engine_row_1based - 1
    test_idx = test_row_1based - 1
    date_idx = date_row_1based - 1
    
    # Auto-detect delimiter
    delimiter = ','
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            sample = f.read(2048)
            # Simple heuristic
            if '\t' in sample:
                delimiter = '\t'
            elif ';' in sample:
                delimiter = ';'
            elif ',' in sample:
                delimiter = ','
    except Exception:
        pass
        
    # Load all rows as raw text list
    rows = []
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f, delimiter=delimiter)
            for r in reader:
                rows.append(r)
    except Exception as e:
        raise ValueError(f"Failed to read file: {e}")
        
    if not rows:
        raise ValueError("File is empty.")
        
    # Standardize column sizes
    max_cols = max(len(r) for r in rows)
    for r in rows:
        r.extend([""] * (max_cols - len(r)))
        
    # Clean and search for columns matching "Take" in test_idx (Row 3)
    # Fallback to engine_idx (Row 2) if not found in test_idx
    keyword_row = test_idx
    date_row = date_idx
    engine_row = engine_idx
    
    use_fallback = True
    # Check if keyword 'Take' exists in Row 3 (keyword_row)
    if len(rows) > keyword_row:
        for val in rows[keyword_row]:
            if "TAKE" in str(val).upper():
                use_fallback = False
                break
                
    if use_fallback and len(rows) > engine_row:
        # Check if keyword 'Take' exists in Row 2 (engine_row)
        for val in rows[engine_row]:
            if "TAKE" in str(val).upper():
                keyword_row = engine_row
                date_row = test_idx # Fallback date row is row 3
                engine_row = -1     # Signal engine is from parent folder
                break
                
    # Gather matching test columns
    test_columns = []
    if len(rows) > keyword_row:
        for c_idx, val in enumerate(rows[keyword_row]):
            if c_idx >= 2 and "TAKE" in str(val).upper():  # Start checking data from Column C (index 2)
                test_columns.append(c_idx)
                
    if not test_columns:
        return [] # No matching test columns
        
    # Determine Engine Name
    parent_folder = os.path.basename(os.path.dirname(file_path))
    
    # Extract Parameter Codes in Column A (index 0)
    param_positions = {} # Param Name -> list of row indices matching this param code
    for r_idx, r in enumerate(rows):
        if r_idx >= 5: # Values start from Excel Row 6
            param_code = str(r[0]).strip()
            if param_code:
                param_positions[param_code] = r_idx
                
    parsed_records = []
    
    for col in test_columns:
        # Engine name
        if custom_engine_name:
            engine_name = custom_engine_name
        elif engine_row != -1 and len(rows) > engine_row:
            engine_name = str(rows[engine_row][col]).strip()
        else:
            engine_name = parent_folder
            
        # Scope / Description
        scope = ""
        if len(rows) > 4: # Excel Row 5 (index 4)
            scope = str(rows[4][col]).strip()
            
        # Date & Time
        date_val = ""
        if len(rows) > date_row:
            raw_date = str(rows[date_row][col]).strip()
            # Try formatting date or use raw string
            date_val = raw_date
            
        # Extract active parameter values
        record = {
            "Engine": engine_name,
            "Date_tested": date_val,
            "Perf. Point": scope
        }
        
        for p in params:
            # We look for simple active params (no formulas)
            if "=" not in p:
                row_idx = param_positions.get(p)
                if row_idx is not None and len(rows[row_idx]) > col:
                    raw_val = rows[row_idx][col]
                    record[p] = clean_numeric(raw_val)
                else:
                    record[p] = np.nan
                    
        parsed_records.append(record)
        
    return parsed_records

def load_database(file_path, sheet_name="Sheet1"):
    if not os.path.exists(file_path):
        return pd.DataFrame()
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        # Convert date column to datetime
        if "Date_tested" in df.columns:
            df["Date_tested"] = pd.to_datetime(df["Date_tested"], errors='coerce')
            df["Date_tested"] = df["Date_tested"].dt.date
        return df
    except Exception:
        return pd.DataFrame()

def save_database(df, file_path, params, sheet_name="Sheet1"):
    """
    Saves the dataframe to Excel.
    Writes calculated columns as Excel formula strings.
    """
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    # Recalculate all formulas in Python first to ensure we write evaluated float values to cells
    # This solves the openpyxl caching problem for Pandas loading
    df_eval = df.copy()
    
    for p in params:
        if "=" in p:
            name, formula = p.split("=", 1)
            # Strip extra '=' in formula definitions like 'W2K3_calc==MAX(...)'
            if formula.startswith("="):
                formula = formula[1:]
            df_eval[name] = evaluate_formula_in_python(df_eval, formula)
            
    # Write to Excel using openpyxl
    wb = load_workbook(file_path) if os.path.exists(file_path) else None
    if wb is None:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
    else:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.delete_rows(1, ws.max_row+1)
        else:
            ws = wb.create_sheet(sheet_name)
            
    # Set headers
    headers = list(df_eval.columns)
    ws.append(headers)
    
    # We write cells. If it's a formula column, we write the evaluated value first
    # AND write the Excel formula string in the cell!
    # A standard Excel formula looks like: '=Formula'
    formula_mappings = {}
    for p in params:
        if "=" in p:
            name, formula = p.split("=", 1)
            if formula.startswith("="):
                formula = formula[1:]
            formula_mappings[name] = formula
            
    # Add rows
    for r_idx, row in df_eval.iterrows():
        row_vals = []
        for col_name in headers:
            val = row[col_name]
            if col_name in formula_mappings:
                # We write the Excel formula! We need to adjust row indexes for table format
                # Structured formulas like =[@[FPR]]*1.01 work automatically inside Excel tables.
                # However, for regular cells we write the Excel formula with correct row index (Excel rows are 1-based, headers are Row 1, data starts at Row 2)
                row_num = r_idx + 2
                excel_formula = formula_mappings[col_name]
                # If formula uses structured table references, e.g. [@[FPR]]
                # we can write it directly as-is!
                row_vals.append(f"={excel_formula}")
            else:
                if pd.isna(val):
                    row_vals.append(None)
                else:
                    row_vals.append(val)
        ws.append(row_vals)
        
    wb.save(file_path)
    wb.close()
