import json
import os
import threading
import uuid
from datetime import datetime
from tkinter import Tk, filedialog

import dash
import dash_ag_grid as dag
import dash_mantine_components as dmc
import openpyxl
from dash import Input, Output, State, callback, dcc, html

from lookup.loopup import (
    add_path_registry_entry,
    get_latest_paths,
    ingest_new_runs,
    preview_run_file,
    read_paths_registry,
    retroactive_parameter_update,
)


dash.register_page(__name__, path="/")

_BASE_DIR = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
_LOOKUP_DIR = os.path.join(_BASE_DIR, "lookup")

_JOB_LOCK = threading.Lock()
_JOBS = {}


def _resolve_registry_path(custom_path=None):
    if custom_path:
        return os.path.normpath(custom_path)

    xlsx_path = os.path.join(_LOOKUP_DIR, "paths.xlsx")
    txt_path = os.path.join(_LOOKUP_DIR, "paths.txt")
    return xlsx_path if os.path.exists(xlsx_path) else txt_path


def _resolve_params_path():
    return os.path.join(_LOOKUP_DIR, "params.txt")


def _resolve_master_excel_from_config():
    config_path = os.path.join(_BASE_DIR, "config.json")
    with open(config_path, "r", encoding="utf-8") as f:
        cfg = json.load(f)

    master_path = cfg.get("data_source", {}).get("file_path")
    if not master_path:
        raise ValueError("config.json is missing data_source.file_path")

    master_path = str(master_path).strip()
    if not os.path.isabs(master_path):
        master_path = os.path.normpath(os.path.join(_BASE_DIR, master_path))
    return master_path


def _load_all_params():
    params_path = _resolve_params_path()
    if not os.path.exists(params_path):
        return []

    params = []
    with open(params_path, "r", encoding="utf-8-sig") as f:
        for raw in f:
            line = raw.strip()
            if not line:
                continue
            params.append(line)
            if "=" in line:
                params.append(line.split("=", 1)[0].strip())

    # De-duplicate while preserving order
    seen = set()
    unique = []
    for p in params:
        if p in seen:
            continue
        seen.add(p)
        unique.append(p)
    return unique


def _create_job(kind):
    job_id = str(uuid.uuid4())
    with _JOB_LOCK:
        _JOBS[job_id] = {
            "kind": kind,
            "status": "running",
            "progress": 0,
            "logs": [f"[{datetime.now().strftime('%H:%M:%S')}] Started {kind} job"],
            "error": "",
        }
    return job_id


def _append_job_log(job_id, message):
    with _JOB_LOCK:
        if job_id not in _JOBS:
            return
        stamp = datetime.now().strftime("%H:%M:%S")
        _JOBS[job_id]["logs"].append(f"[{stamp}] {message}")
        _JOBS[job_id]["logs"] = _JOBS[job_id]["logs"][-300:]


def _set_job_progress(job_id, value):
    with _JOB_LOCK:
        if job_id in _JOBS:
            _JOBS[job_id]["progress"] = max(0, min(100, int(value)))


def _finish_job(job_id):
    with _JOB_LOCK:
        if job_id in _JOBS:
            _JOBS[job_id]["status"] = "completed"
            _JOBS[job_id]["progress"] = 100


def _fail_job(job_id, err):
    with _JOB_LOCK:
        if job_id in _JOBS:
            _JOBS[job_id]["status"] = "failed"
            _JOBS[job_id]["error"] = str(err)


def _get_job_snapshot(job_id):
    with _JOB_LOCK:
        job = _JOBS.get(job_id)
        if not job:
            return None
        return {
            "status": job.get("status"),
            "progress": job.get("progress", 0),
            "logs": list(job.get("logs", [])),
            "error": job.get("error", ""),
        }


def _render_preview(preview):
    if not preview or not preview.get("ok"):
        return dmc.Alert(preview.get("message", "Preview unavailable"), color="red", variant="light")

    tests = preview.get("take_tests", [])
    dates = preview.get("tested_dates", [])

    items = []
    for idx, test_name in enumerate(tests):
        date_label = dates[idx] if idx < len(dates) else ""
        items.append(html.Li(f"{test_name} {('(' + date_label + ')') if date_label else ''}"))

    return dmc.Paper(
        withBorder=True,
        p="sm",
        children=[
            dmc.Text(f"Path: {preview.get('path', '')}", size="sm"),
            dmc.Text(
                f"Engine: {preview.get('engine') or 'Not detected'}",
                fw=600,
                mt="xs",
            ),
            dmc.Text(f"Matching take-off tests: {len(tests)}", mt="xs"),
            html.Ul(items if items else [html.Li("No matching test columns found")]),
        ],
    )


def _run_add_file_job(job_id, run_path, selected_engine=None):
    try:
        master_excel = _resolve_master_excel_from_config()
        all_params = _load_all_params()
        engine_value = str(selected_engine or "").strip()

        _append_job_log(job_id, f"Target database: {master_excel}")
        _append_job_log(job_id, f"Source run file: {run_path}")
        if engine_value:
            _append_job_log(job_id, f"Engine override: {engine_value}")

        overrides = {run_path: engine_value} if engine_value else None
        for step in ingest_new_runs([run_path], all_params, master_excel, engine_overrides=overrides):
            progress = int((step.get("progress", 0) / max(step.get("total", 1), 1)) * 100)
            _set_job_progress(job_id, progress)
            _append_job_log(job_id, step.get("message", "Running..."))

        reg_result = add_path_registry_entry(run_path, engine_name=engine_value)
        _append_job_log(job_id, f"Path registry: {reg_result.get('message', 'done')}")
        _finish_job(job_id)
    except Exception as err:
        _append_job_log(job_id, f"ERROR: {err}")
        _fail_job(job_id, err)


def _run_add_parameter_job(job_id, param_text, paths_registry):
    try:
        params_path = _resolve_params_path()
        master_excel = _resolve_master_excel_from_config()
        registry_path = _resolve_registry_path(paths_registry)

        _append_job_log(job_id, f"New parameter: {param_text}")
        _append_job_log(job_id, f"Paths source: {registry_path}")

        for step in retroactive_parameter_update(param_text, registry_path, params_path, master_excel):
            progress = int((step.get("progress", 0) / max(step.get("total", 1), 1)) * 100)
            _set_job_progress(job_id, progress)
            _append_job_log(job_id, step.get("message", "Running..."))

        _finish_job(job_id)
    except Exception as err:
        _append_job_log(job_id, f"ERROR: {err}")
        _fail_job(job_id, err)


def _prepare_master_for_rescan(master_excel):
    wb = openpyxl.load_workbook(master_excel)
    ws = wb.active

    # Rescan contract: first 3 columns are always Engine, Date_Tested, Perf. Point.
    ws.cell(row=1, column=1, value="Engine")
    ws.cell(row=1, column=2, value="Date_Tested")
    ws.cell(row=1, column=3, value="Perf. Point")

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    wb.save(master_excel)


def _run_rescan_all_job(job_id):
    try:
        master_excel = _resolve_master_excel_from_config()
        paths_df = read_paths_registry(latest_first=False)
        all_params = _load_all_params()

        if paths_df.empty or "Path" not in paths_df.columns:
            raise ValueError("No paths available for rescan.")

        candidate_paths = [str(p).strip() for p in paths_df["Path"].tolist() if str(p).strip()]
        valid_paths = [p for p in candidate_paths if os.path.exists(p)]
        missing_paths = [p for p in candidate_paths if not os.path.exists(p)]

        if not valid_paths:
            raise ValueError("No valid file paths found for rescan.")

        _append_job_log(job_id, f"Master DB: {master_excel}")
        _append_job_log(job_id, f"Params loaded: {len(all_params)}")
        _append_job_log(job_id, f"Paths to scan: {len(valid_paths)}")
        if missing_paths:
            _append_job_log(job_id, f"Skipped missing paths: {len(missing_paths)}")

        _set_job_progress(job_id, 5)
        _append_job_log(job_id, "Preparing master workbook for full rebuild...")
        _prepare_master_for_rescan(master_excel)

        engine_overrides = {}
        if "Engine" in paths_df.columns:
            for _, row in paths_df.iterrows():
                path_val = str(row.get("Path", "")).strip()
                eng_val = str(row.get("Engine", "")).strip()
                if path_val and eng_val:
                    engine_overrides[path_val] = eng_val

        for step in ingest_new_runs(valid_paths, all_params, master_excel, engine_overrides=engine_overrides):
            progress = int((step.get("progress", 0) / max(step.get("total", 1), 1)) * 100)
            _set_job_progress(job_id, progress)
            _append_job_log(job_id, step.get("message", "Running..."))

        _finish_job(job_id)
    except Exception as err:
        _append_job_log(job_id, f"ERROR: {err}")
        _fail_job(job_id, err)


def layout():
    return html.Div(
        [
            dcc.Interval(id="lookup-initial-load", interval=200, n_intervals=0, max_intervals=1),
            dcc.Interval(id="lookup-job-poller", interval=800, n_intervals=0),
            dcc.Store(id="add-file-preview-store", data=None),
            dcc.Store(id="add-param-preview-store", data=None),
            dcc.Store(id="rescan-preview-store", data=None),
            dcc.Store(id="add-file-job-id", data=None),
            dcc.Store(id="add-param-job-id", data=None),
            dcc.Store(id="rescan-job-id", data=None),
            dmc.Space(h=52),
            dmc.Title("Lookup Operations", order=2),
            dmc.Text("Run file ingestion and parameter workflows with preview, confirmation, progress bars, and logs.", c="dimmed", mb="md"),
            dmc.Grid(
                gutter="md",
                children=[
                    dmc.GridCol(
                        span=12,
                        children=dmc.Paper(
                            withBorder=True,
                            p="md",
                            children=[
                                dmc.Group(
                                    [
                                        dmc.Title("Latest Paths", order=4),
                                        dmc.Group(
                                            [
                                                dmc.Button("Refresh", id="lookup-refresh-btn", variant="outline"),
                                                dmc.Button("Show Actions", id="toggle-actions-btn", variant="light"),
                                            ],
                                            gap="xs",
                                        ),
                                    ],
                                    justify="space-between",
                                    mb="sm",
                                ),
                                dmc.Text(id="lookup-registry-source", size="sm", c="dimmed", mb="xs"),
                                dag.AgGrid(
                                    id="lookup-paths-grid",
                                    columnDefs=[
                                        {"headerName": "Path", "field": "Path", "flex": 3, "filter": True},
                                        {"headerName": "Engine", "field": "Engine", "flex": 1.2, "filter": True},
                                        {"headerName": "Date Tested", "field": "Date_Tested", "flex": 1.2, "filter": True},
                                        {"headerName": "Date Added", "field": "Date_Added", "flex": 1.2, "sort": "desc", "filter": True},
                                        {"headerName": "Added By", "field": "Added_By", "flex": 1, "filter": True},
                                    ],
                                    defaultColDef={"sortable": True, "resizable": True},
                                    rowData=[],
                                    dashGridOptions={"pagination": True, "paginationPageSize": 20, "animateRows": True},
                                    className="ag-theme-alpine",
                                    style={"height": "50vh", "width": "100%"},
                                ),
                            ],
                        ),
                    ),
                    dmc.GridCol(
                        span=12,
                        children=html.Div(
                            id="lookup-actions-container",
                            style={"display": "none"},
                            children=dmc.Grid(
                                gutter="md",
                                children=[
                                    dmc.GridCol(
                                        span=12,
                                        children=dmc.Paper(
                                            withBorder=True,
                                            p="md",
                                            children=[
                                                dmc.Group(
                                                    [
                                                        dmc.Title("Add File", order=4),
                                                        dmc.Button("Browse File", id="add-file-browse-btn", variant="light"),
                                                    ],
                                                    justify="space-between",
                                                    mb="sm",
                                                ),
                                                dmc.TextInput(
                                                    id="add-file-path",
                                                    placeholder=r"Paste full file path (required for script run), e.g. C:\\Runs\\WW284A.txt",
                                                    mb="sm",
                                                ),
                                                dmc.TextInput(
                                                    id="add-file-engine",
                                                    placeholder="Engine name (auto-filled if detected; required if not detected)",
                                                    mb="sm",
                                                ),
                                                dmc.Group(
                                                    [
                                                        dmc.Button("Preview File", id="add-file-preview-btn", color="blue"),
                                                        dmc.Button("Confirm and Run", id="add-file-confirm-btn", color="green"),
                                                    ],
                                                    mb="sm",
                                                ),
                                                dmc.Text(id="add-file-upload-note", c="dimmed", size="sm", mb="xs", children="Use Browse File to select an absolute path."),
                                                html.Div(id="add-file-preview", className="mb-2"),
                                                html.Div(id="add-file-status"),
                                                dmc.Progress(id="add-file-progress", value=0, mb="xs"),
                                                dmc.Text(id="add-file-progress-text", size="sm", c="dimmed", mb="xs", children="Idle"),
                                                html.Pre(id="add-file-logs", style={"maxHeight": "200px", "overflowY": "auto", "background": "#f8f9fa", "padding": "10px"}),
                                                html.Div(id="add-file-runtime-status"),
                                            ],
                                        ),
                                    ),
                                    dmc.GridCol(
                                        span=12,
                                        children=dmc.Paper(
                                            withBorder=True,
                                            p="md",
                                            children=[
                                                dmc.Title("Add Parameter", order=4, mb="sm"),
                                                dmc.TextInput(
                                                    id="add-param-input",
                                                    placeholder=r"Enter parameter, e.g. A01111 or NEW_PARAM=([A]+[B])/2",
                                                    mb="sm",
                                                ),
                                                dmc.TextInput(
                                                    id="add-param-paths-source",
                                                    placeholder=r"Optional paths source (defaults to lookup\\paths.xlsx or paths.txt)",
                                                    mb="sm",
                                                ),
                                                dmc.Group(
                                                    [
                                                        dmc.Button("Preview Change", id="add-param-preview-btn", color="blue"),
                                                        dmc.Button("Confirm and Run", id="add-param-confirm-btn", color="green"),
                                                    ],
                                                    mb="sm",
                                                ),
                                                html.Div(id="add-param-preview", className="mb-2"),
                                                html.Div(id="add-param-status"),
                                                dmc.Progress(id="add-param-progress", value=0, mb="xs"),
                                                dmc.Text(id="add-param-progress-text", size="sm", c="dimmed", mb="xs", children="Idle"),
                                                html.Pre(id="add-param-logs", style={"maxHeight": "200px", "overflowY": "auto", "background": "#f8f9fa", "padding": "10px"}),
                                                html.Div(id="add-param-runtime-status"),
                                            ],
                                        ),
                                    ),
                                    dmc.GridCol(
                                        span=12,
                                        children=dmc.Paper(
                                            withBorder=True,
                                            p="md",
                                            children=[
                                                dmc.Title("Rescan All Paths", order=4, mb="sm"),
                                                dmc.Text(
                                                    "Rebuilds main database from all paths and all params. This clears existing data rows first.",
                                                    size="sm",
                                                    c="dimmed",
                                                    mb="sm",
                                                ),
                                                dmc.Group(
                                                    [
                                                        dmc.Button("Preview Rescan", id="rescan-preview-btn", color="blue"),
                                                        dmc.Button("Confirm Rescan", id="rescan-confirm-btn", color="red"),
                                                    ],
                                                    mb="sm",
                                                ),
                                                html.Div(id="rescan-preview", className="mb-2"),
                                                html.Div(id="rescan-status"),
                                                dmc.Progress(id="rescan-progress", value=0, mb="xs"),
                                                dmc.Text(id="rescan-progress-text", size="sm", c="dimmed", mb="xs", children="Idle"),
                                                html.Pre(id="rescan-logs", style={"maxHeight": "200px", "overflowY": "auto", "background": "#f8f9fa", "padding": "10px"}),
                                                html.Div(id="rescan-runtime-status"),
                                            ],
                                        ),
                                    ),
                                ],
                            ),
                        ),
                    ),
                ],
            ),
        ],
        style={"marginTop": "20px"},
    )


@callback(
    Output("lookup-actions-container", "style"),
    Output("toggle-actions-btn", "children"),
    Input("toggle-actions-btn", "n_clicks"),
)
def toggle_actions_panel(n_clicks):
    is_open = bool(n_clicks and n_clicks % 2 == 1)
    if is_open:
        return {"display": "block"}, "Hide Actions"
    return {"display": "none"}, "Show Actions"


@callback(
    Output("lookup-paths-grid", "rowData"),
    Output("lookup-registry-source", "children"),
    Input("lookup-initial-load", "n_intervals"),
    Input("lookup-refresh-btn", "n_clicks"),
)
def refresh_lookup_registry(_, _refresh_clicks):
    rows = get_latest_paths(limit=500)
    source_text = f"Registry source: {_resolve_registry_path()}"
    return rows, source_text


@callback(
    Output("add-file-path", "value"),
    Output("add-file-upload-note", "children"),
    Input("add-file-browse-btn", "n_clicks"),
    prevent_initial_call=True,
)
def show_upload_filename(_n_clicks):
    try:
        root = Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        selected = filedialog.askopenfilename(title="Select test file")
        root.destroy()
    except Exception as err:
        return dash.no_update, f"Could not open native picker: {err}"

    if not selected:
        return dash.no_update, "File selection cancelled."

    return os.path.normpath(selected), "Full absolute path selected from native file dialog."


@callback(
    Output("add-file-preview-store", "data"),
    Output("add-file-preview", "children"),
    Output("add-file-status", "children"),
    Output("add-file-engine", "value"),
    Input("add-file-preview-btn", "n_clicks"),
    State("add-file-path", "value"),
    prevent_initial_call=True,
)
def preview_add_file(_n_clicks, run_path):
    preview = preview_run_file(run_path)
    if not preview.get("ok"):
        return None, dmc.Alert(preview.get("message", "Preview failed"), color="red", variant="light"), dash.no_update, dash.no_update

    detected_engine = str(preview.get("engine") or "").strip()
    engine_detected = bool(preview.get("engine_detected"))
    status_message = "Preview complete. Confirm to append to main database."
    status_color = "green"
    if not engine_detected:
        status_message = "Engine row not detected. Enter engine name manually before confirm."
        status_color = "orange"

    return (
        {
            "path": preview.get("path"),
            "engine_detected": engine_detected,
            "detected_engine": detected_engine,
        },
        _render_preview(preview),
        dmc.Alert(status_message, color=status_color, variant="light"),
        detected_engine if engine_detected else "",
    )


@callback(
    Output("add-file-job-id", "data"),
    Output("add-file-status", "children", allow_duplicate=True),
    Input("add-file-confirm-btn", "n_clicks"),
    State("add-file-preview-store", "data"),
    State("add-file-path", "value"),
    State("add-file-engine", "value"),
    prevent_initial_call=True,
)
def start_add_file_job(_n_clicks, preview_data, manual_path, manual_engine):
    run_path = (preview_data or {}).get("path") or str(manual_path or "").strip()
    if not run_path:
        return dash.no_update, dmc.Alert("Preview the file first or provide a valid path.", color="red", variant="light")

    detected_engine = str((preview_data or {}).get("detected_engine") or "").strip()
    engine_detected = bool((preview_data or {}).get("engine_detected"))
    selected_engine = str(manual_engine or "").strip() or detected_engine

    if not engine_detected and not selected_engine:
        return dash.no_update, dmc.Alert(
            "Engine row was not detected. Enter an engine name before running.",
            color="red",
            variant="light",
        )

    job_id = _create_job("add_file")
    thread = threading.Thread(target=_run_add_file_job, args=(job_id, run_path, selected_engine), daemon=True)
    thread.start()
    return job_id, dmc.Alert("Add file job started.", color="blue", variant="light")


@callback(
    Output("add-param-preview-store", "data"),
    Output("add-param-preview", "children"),
    Output("add-param-status", "children"),
    Input("add-param-preview-btn", "n_clicks"),
    State("add-param-input", "value"),
    State("add-param-paths-source", "value"),
    prevent_initial_call=True,
)
def preview_add_parameter(_n_clicks, param_text, paths_source):
    param_text = str(param_text or "").strip()
    if not param_text:
        return None, dmc.Alert("Parameter value is required.", color="red", variant="light"), dash.no_update

    registry_path = _resolve_registry_path(paths_source if str(paths_source or "").strip() else None)
    try:
        registry_df = read_paths_registry(paths_registry=registry_path, latest_first=False)
    except Exception as err:
        return None, dmc.Alert(f"Could not read path source: {err}", color="red", variant="light"), dash.no_update

    header = param_text.split("=", 1)[0].strip() if "=" in param_text else param_text
    preview_component = dmc.Paper(
        withBorder=True,
        p="sm",
        children=[
            dmc.Text(f"Parameter header: {header}", fw=600),
            dmc.Text(f"Expression/raw value: {param_text}", size="sm"),
            dmc.Text(f"Path source: {registry_path}", size="sm"),
            dmc.Text(f"Files that will be scanned: {len(registry_df)}", size="sm"),
        ],
    )

    return (
        {"param_text": param_text, "paths_source": registry_path},
        preview_component,
        dmc.Alert("Preview complete. Confirm to run retroactive parameter update.", color="green", variant="light"),
    )


@callback(
    Output("add-param-job-id", "data"),
    Output("add-param-status", "children", allow_duplicate=True),
    Input("add-param-confirm-btn", "n_clicks"),
    State("add-param-preview-store", "data"),
    State("add-param-input", "value"),
    State("add-param-paths-source", "value"),
    prevent_initial_call=True,
)
def start_add_parameter_job(_n_clicks, preview_data, param_text, paths_source):
    payload = preview_data or {}
    final_param = payload.get("param_text") or str(param_text or "").strip()
    final_source = payload.get("paths_source") or str(paths_source or "").strip() or None

    if not final_param:
        return dash.no_update, dmc.Alert("Preview parameter first or provide a valid parameter value.", color="red", variant="light")

    job_id = _create_job("add_parameter")
    thread = threading.Thread(target=_run_add_parameter_job, args=(job_id, final_param, final_source), daemon=True)
    thread.start()
    return job_id, dmc.Alert("Add parameter job started.", color="blue", variant="light")


@callback(
    Output("rescan-preview-store", "data"),
    Output("rescan-preview", "children"),
    Output("rescan-status", "children"),
    Input("rescan-preview-btn", "n_clicks"),
    prevent_initial_call=True,
)
def preview_rescan_all(_n_clicks):
    try:
        paths_df = read_paths_registry(latest_first=False)
        all_params = _load_all_params()

        if paths_df.empty or "Path" not in paths_df.columns:
            return None, dmc.Alert("No paths available for rescan.", color="red", variant="light"), dash.no_update

        candidate_paths = [str(p).strip() for p in paths_df["Path"].tolist() if str(p).strip()]
        valid_paths = [p for p in candidate_paths if os.path.exists(p)]
        missing_paths = [p for p in candidate_paths if not os.path.exists(p)]

        preview = dmc.Paper(
            withBorder=True,
            p="sm",
            children=[
                dmc.Text(f"Total paths in registry: {len(candidate_paths)}", size="sm"),
                dmc.Text(f"Valid paths to process: {len(valid_paths)}", size="sm"),
                dmc.Text(f"Missing paths to skip: {len(missing_paths)}", size="sm"),
                dmc.Text(f"Params loaded: {len(all_params)}", size="sm"),
                dmc.Text("Master headers enforced: Engine, Date_Tested, Perf. Point", size="sm"),
            ],
        )

        return (
            {"ready": True, "valid_paths": len(valid_paths), "params": len(all_params)},
            preview,
            dmc.Alert("Preview complete. Confirm to run full rescan.", color="orange", variant="light"),
        )
    except Exception as err:
        return None, dmc.Alert(f"Rescan preview failed: {err}", color="red", variant="light"), dash.no_update


@callback(
    Output("rescan-job-id", "data"),
    Output("rescan-status", "children", allow_duplicate=True),
    Input("rescan-confirm-btn", "n_clicks"),
    State("rescan-preview-store", "data"),
    prevent_initial_call=True,
)
def start_rescan_all_job(_n_clicks, preview_data):
    if not preview_data or not preview_data.get("ready"):
        return dash.no_update, dmc.Alert("Run Preview Rescan before confirmation.", color="red", variant="light")
    if int(preview_data.get("valid_paths", 0)) <= 0:
        return dash.no_update, dmc.Alert("No valid paths to rescan.", color="red", variant="light")

    job_id = _create_job("rescan_all")
    thread = threading.Thread(target=_run_rescan_all_job, args=(job_id,), daemon=True)
    thread.start()
    return job_id, dmc.Alert("Rescan job started.", color="blue", variant="light")


def _poll_job_to_ui(job_id):
    if not job_id:
        return 0, "Idle", "", dash.no_update

    snap = _get_job_snapshot(job_id)
    if not snap:
        return 0, "Idle", "", dmc.Alert("Job not found.", color="red", variant="light")

    progress = snap.get("progress", 0)
    status = snap.get("status", "running")
    logs = "\n".join(snap.get("logs", []))

    if status == "running":
        runtime = dmc.Alert("Running...", color="blue", variant="light")
        return progress, f"{progress}%", logs, runtime

    if status == "completed":
        runtime = dmc.Alert("Completed.", color="green", variant="light")
        return progress, "100%", logs, runtime

    runtime = dmc.Alert(f"Failed: {snap.get('error', 'Unknown error')}", color="red", variant="light")
    return progress, f"{progress}%", logs, runtime


@callback(
    Output("add-file-progress", "value"),
    Output("add-file-progress-text", "children"),
    Output("add-file-logs", "children"),
    Output("add-file-runtime-status", "children"),
    Input("lookup-job-poller", "n_intervals"),
    State("add-file-job-id", "data"),
)
def poll_add_file_job(_n, job_id):
    return _poll_job_to_ui(job_id)


@callback(
    Output("add-param-progress", "value"),
    Output("add-param-progress-text", "children"),
    Output("add-param-logs", "children"),
    Output("add-param-runtime-status", "children"),
    Input("lookup-job-poller", "n_intervals"),
    State("add-param-job-id", "data"),
)
def poll_add_parameter_job(_n, job_id):
    return _poll_job_to_ui(job_id)


@callback(
    Output("rescan-progress", "value"),
    Output("rescan-progress-text", "children"),
    Output("rescan-logs", "children"),
    Output("rescan-runtime-status", "children"),
    Input("lookup-job-poller", "n_intervals"),
    State("rescan-job-id", "data"),
)
def poll_rescan_job(_n, job_id):
    return _poll_job_to_ui(job_id)
